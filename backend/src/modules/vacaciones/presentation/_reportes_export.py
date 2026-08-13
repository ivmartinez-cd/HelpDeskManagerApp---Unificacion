"""Exportadores Excel/PDF del reporte de vacaciones.

Paridad con `report.controller.ts` legacy: mismos nombres de archivo, mismas
hojas/columnas del Excel y mismo layout de texto del PDF (título + fecha de
generación + listas por empleado y por sector). Única diferencia deliberada:
"Departamento" pasa a llamarse "Sector" (vocabulario de esta app).
"""

import io
from datetime import datetime
from zoneinfo import ZoneInfo

from fastapi.responses import StreamingResponse
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from src.modules.vacaciones.application.dtos.reporte_dtos import ReporteVacacionesDTO

_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _attachment(buf: io.BytesIO, media_type: str, filename: str) -> StreamingResponse:
    return StreamingResponse(
        buf,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def export_excel(reporte: ReporteVacacionesDTO) -> StreamingResponse:
    wb = Workbook()
    hoja_empleados = wb.active
    assert hoja_empleados is not None  # openpyxl siempre crea la hoja inicial
    _hoja_empleados(hoja_empleados, reporte)
    _hoja_sectores(wb.create_sheet("Por sector"), reporte)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _attachment(buf, _XLSX_MEDIA_TYPE, "reporte-vacaciones.xlsx")


def _hoja_empleados(ws: Worksheet, reporte: ReporteVacacionesDTO) -> None:
    ws.title = "Por empleado"
    headers = [
        ("Empleado", 28),
        ("Sector", 20),
        ("Cargo", 24),
        ("Días anuales", 14),
        ("Consumidos", 14),
        ("Pendientes", 14),
        ("Disponibles", 14),
    ]
    _encabezado(ws, headers)
    for f in reporte.por_empleado:
        ws.append(
            [f.nombre, f.sector_nombre, f.cargo_nombre, f.annual, f.used, f.pending, f.available]
        )


def _hoja_sectores(ws: Worksheet, reporte: ReporteVacacionesDTO) -> None:
    headers = [
        ("Sector", 24),
        ("Empleados", 14),
        ("Días anuales", 14),
        ("Consumidos", 14),
        ("Disponibles", 14),
    ]
    _encabezado(ws, headers)
    for f in reporte.por_sector:
        ws.append([f.nombre, f.empleados, f.annual, f.used, f.available])


def _encabezado(ws: Worksheet, headers: list[tuple[str, int]]) -> None:
    ws.append([h for h, _ in headers])
    for idx, (_, width) in enumerate(headers):
        ws.column_dimensions[chr(ord("A") + idx)].width = width
    for cell in ws[1]:
        cell.font = Font(bold=True)


def export_pdf(reporte: ReporteVacacionesDTO, *, timezone: str) -> StreamingResponse:
    pdf = FPDF(format="A4")
    pdf.set_margins(14, 14)
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()
    _pdf_titulo(pdf, timezone)
    _pdf_seccion(
        pdf,
        "Vacaciones por empleado",
        [
            f"{f.nombre} - {f.sector_nombre} | Anuales: {f.annual}  Consumidos: {f.used}"
            f"  Pendientes: {f.pending}  Disponibles: {f.available}"
            for f in reporte.por_empleado
        ],
    )
    pdf.ln(6)
    _pdf_seccion(
        pdf,
        "Vacaciones por sector",
        [
            f"{f.nombre} - Empleados: {f.empleados} | Anuales: {f.annual}"
            f"  Consumidos: {f.used}  Disponibles: {f.available}"
            for f in reporte.por_sector
        ],
    )
    buf = io.BytesIO(bytes(pdf.output()))
    return _attachment(buf, "application/pdf", "reporte-vacaciones.pdf")


def _pdf_titulo(pdf: FPDF, timezone: str) -> None:
    ahora = datetime.now(ZoneInfo(timezone))
    pdf.set_font("helvetica", size=20)
    pdf.set_text_color(30, 41, 59)
    pdf.cell(0, 10, "Reporte de Vacaciones", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", size=10)
    pdf.set_text_color(100, 116, 139)
    generado = f"Generado el {ahora.strftime('%d/%m/%Y %H:%M')}"
    pdf.cell(0, 6, generado, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)


def _pdf_seccion(pdf: FPDF, titulo: str, lineas: list[str]) -> None:
    pdf.set_font("helvetica", size=14)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(0, 8, titulo, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    pdf.set_font("helvetica", size=9)
    pdf.set_text_color(51, 65, 85)
    for linea in lineas:
        pdf.multi_cell(0, 5, _latin1(linea), new_x="LMARGIN", new_y="NEXT")


def _latin1(texto: str) -> str:
    """Las fuentes core de fpdf2 solo cubren latin-1; se degrada lo que no entre."""
    return texto.encode("latin-1", "replace").decode("latin-1")
