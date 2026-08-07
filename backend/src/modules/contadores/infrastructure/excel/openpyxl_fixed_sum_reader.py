from openpyxl import load_workbook

from src.modules.contadores.domain.errors import InvalidCounterWorkbookError, MissingColumnError
from src.modules.contadores.domain.value_objects.fixed_sum_source_row import FixedSumSourceRow

_SERIE_ALIASES = ("SERIE", "Nro Serie")


class OpenpyxlFixedSumReader:
    """Lee el Excel de Suma Fija: encabezados en la fila 1 (sin fila de
    título, a diferencia del de Proyección)."""

    def read(self, file_path: str) -> list[FixedSumSourceRow]:
        rows = _load_rows(file_path)
        if not rows:
            return []
        columns = {str(c).strip(): i for i, c in enumerate(rows[0])}
        serie_idx = _find_column(columns, _SERIE_ALIASES)
        estado_idx = _find_column(columns, ("Estado",))
        contador_idx = _find_column(columns, ("Cdor Actual",))
        return [_parse_row(r, serie_idx, estado_idx, contador_idx) for r in rows[1:]]


def _load_rows(file_path: str) -> list[tuple[object, ...]]:
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        raise InvalidCounterWorkbookError(f"No se pudo abrir el archivo Excel: {exc}") from exc
    ws = wb[wb.sheetnames[0]]
    rows: list[tuple[object, ...]] = [tuple(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _find_column(columns: dict[str, int], aliases: tuple[str, ...]) -> int:
    for alias in aliases:
        if alias in columns:
            return columns[alias]
    raise MissingColumnError(aliases[0])


def _parse_row(
    row: tuple[object, ...], serie_idx: int, estado_idx: int, contador_idx: int
) -> FixedSumSourceRow:
    contador_raw = row[contador_idx]
    contador = int(float(contador_raw)) if contador_raw is not None else 0  # type: ignore[arg-type]
    return FixedSumSourceRow(
        serie=str(row[serie_idx] or ""),
        estado=str(row[estado_idx] or ""),
        contador_actual=contador,
    )
