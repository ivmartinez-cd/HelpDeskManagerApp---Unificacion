import csv
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.modules.contadores.domain.repositories.projection_report_writer import (
    ProjectionReportPaths,
)
from src.modules.contadores.domain.value_objects.audit_entry import AuditEntry
from src.modules.contadores.domain.value_objects.counter_projection_result import (
    CounterProjectionResult,
)
from src.modules.contadores.domain.value_objects.projection_summary import ProjectionSummary
from src.modules.contadores.domain.value_objects.siges_row import SigesRow

_SIGES_FIELDNAMES = (
    "SERIE",
    "FECHA",
    "TIPO",
    "CLASE_10",
    "CONTADOR_10",
    "CLASE_20",
    "CONTADOR_20",
    "MOTIVO",
    "OBSERVACIONES",
)


class OpenpyxlProjectionReportWriter:
    """Escribe el Excel de resultados (hojas Proyección/Auditoría/Resumen) y
    el CSV de SiGes, en el formato que ya consume el proceso de facturación
    (`;` como separador, CRLF, UTF-8 sin BOM — igual que la app vieja)."""

    def write(
        self,
        *,
        results: list[CounterProjectionResult],
        audit: list[AuditEntry],
        siges_rows: list[SigesRow],
        summary: ProjectionSummary,
        fecha_toma: date,
        source_filename: str,
        output_dir: str,
    ) -> ProjectionReportPaths:
        out = Path(output_dir)
        out.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        excel_path = out / f"Reporte_Proyeccion_{stamp}.xlsx"
        _write_workbook(excel_path, results, audit, summary, fecha_toma, source_filename)

        csv_path = out / f"import_siges_proyeccion_{stamp}.csv"
        _write_siges_csv(csv_path, siges_rows)

        return ProjectionReportPaths(excel_path=str(excel_path), siges_csv_path=str(csv_path))


def _write_workbook(
    path: Path,
    results: list[CounterProjectionResult],
    audit: list[AuditEntry],
    summary: ProjectionSummary,
    fecha_toma: date,
    source_filename: str,
) -> None:
    wb = Workbook()
    proyeccion_sheet = wb.active
    assert proyeccion_sheet is not None  # Workbook() siempre crea una hoja activa
    proyeccion_sheet.title = "Proyeccion"
    _write_results_sheet(proyeccion_sheet, results)
    _write_audit_sheet(wb.create_sheet("Auditoria"), audit)
    _write_summary_sheet(wb.create_sheet("Resumen"), summary, fecha_toma, source_filename)
    wb.save(path)


def _write_results_sheet(ws: Worksheet, results: list[CounterProjectionResult]) -> None:
    headers = (
        "Nro Serie",
        "Clase",
        "Articulo",
        "Sector",
        "Fecha Ultima Lectura",
        "Contador Base",
        "Dias Proyectados",
        "Consumo Diario Promedio",
        "Paginas Sumadas",
        "Fecha Toma",
        "Contador Proyectado",
        "Metodo",
        "Observaciones",
    )
    ws.append(headers)
    for r in results:
        ws.append(
            (
                r.serie,
                r.clase,
                r.articulo,
                r.sector,
                r.fecha_lectura,
                r.contador_base,
                r.dias_proyectados,
                r.consumo_diario_promedio,
                r.paginas_sumadas,
                r.fecha_toma,
                r.contador_proyectado,
                r.metodo,
                r.observaciones,
            )
        )


def _write_audit_sheet(ws: Worksheet, audit: list[AuditEntry]) -> None:
    ws.append(("Nro Serie", "Clase", "Fecha", "Contador", "Tipo Contador", "Usado", "Motivo"))
    for a in audit:
        ws.append((a.serie, a.clase, a.fecha, a.contador, a.tipo_contador, a.usado, a.motivo))


def _write_summary_sheet(
    ws: Worksheet, summary: ProjectionSummary, fecha_toma: date, source_filename: str
) -> None:
    ws.append(("Indicador", "Valor"))
    rows = (
        ("Archivo", source_filename),
        ("Fecha de toma", fecha_toma),
        ("Total procesados", summary.total),
        ("REAL", summary.reales),
        ("PROYECTADO", summary.proyectados),
        ("SIN DATOS", summary.sin_datos),
        ("Dias proyectados - mediana", summary.dias_mediana),
        ("Dias proyectados - max", summary.dias_max),
        ("Consumo diario - mediana", summary.consumo_mediana),
        ("Consumo diario - max", summary.consumo_max),
    )
    for row in rows:
        ws.append(row)


def _write_siges_csv(path: Path, rows: list[SigesRow]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";", lineterminator="\r\n")
        writer.writerow(_SIGES_FIELDNAMES)
        for r in rows:
            writer.writerow(
                (
                    r.serie,
                    r.fecha.strftime("%d/%m/%Y"),
                    r.tipo,
                    r.clase_10,
                    r.contador_10,
                    r.clase_20,
                    r.contador_20,
                    r.motivo,
                    r.observaciones,
                )
            )
