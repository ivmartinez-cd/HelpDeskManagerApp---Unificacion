from dataclasses import dataclass
from datetime import date, datetime

from openpyxl import load_workbook

from src.modules.contadores.domain.errors import (
    EmptyCounterWorkbookError,
    InvalidCounterWorkbookError,
)
from src.modules.contadores.domain.value_objects.counter_reading import CounterReading
from src.modules.contadores.domain.value_objects.device_readings import DeviceReadings

_REQUIRED_COLUMNS = (
    "Articulo",
    "Nro Serie",
    "Sector",
    "Fecha",
    "Tipo Contador",
    "Clase",
    "Contador",
)


@dataclass(frozen=True, slots=True)
class _ParsedRow:
    articulo: str
    serie: str
    sector: str
    fecha: date
    tipo_contador: str
    clase: str
    contador: int


class OpenpyxlCounterWorkbookReader:
    """Lee el Excel de lecturas de contadores. Formato SSRS: fila 1 es un
    título libre, fila 2 son los encabezados reales — igual que la app vieja
    (`pd.read_excel(header=1)`)."""

    def read(self, file_path: str) -> list[DeviceReadings]:
        rows = _load_rows(file_path)
        columns = {str(c).strip(): i for i, c in enumerate(rows[0])} if rows else {}
        _validate_columns(columns)

        grouped: dict[tuple[str, str], list[_ParsedRow]] = {}
        for raw in rows[1:]:
            parsed = _parse_row(raw, columns)
            if parsed is None:
                continue
            grouped.setdefault((parsed.serie, parsed.clase), []).append(parsed)

        if not grouped:
            raise EmptyCounterWorkbookError
        return [_build_device(serie, clase, items) for (serie, clase), items in grouped.items()]


def _load_rows(file_path: str) -> list[tuple[object, ...]]:
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        raise InvalidCounterWorkbookError(f"No se pudo abrir el archivo Excel: {exc}") from exc
    ws = wb[wb.sheetnames[0]]
    rows: list[tuple[object, ...]] = [
        tuple(row) for row in ws.iter_rows(min_row=2, values_only=True)
    ]
    wb.close()
    return rows


def _validate_columns(columns: dict[str, int]) -> None:
    faltantes = [c for c in _REQUIRED_COLUMNS if c not in columns]
    if faltantes:
        raise InvalidCounterWorkbookError(f"Faltan columnas requeridas: {', '.join(faltantes)}")


def _parse_row(raw: tuple[object, ...], columns: dict[str, int]) -> _ParsedRow | None:
    fecha = _parse_fecha(raw[columns["Fecha"]])
    contador = _parse_contador(raw[columns["Contador"]])
    if fecha is None or contador is None:
        return None
    return _ParsedRow(
        articulo=str(raw[columns["Articulo"]] or ""),
        serie=str(raw[columns["Nro Serie"]]),
        sector=str(raw[columns["Sector"]] or ""),
        fecha=fecha,
        tipo_contador=str(raw[columns["Tipo Contador"]] or ""),
        clase=str(raw[columns["Clase"]]),
        contador=contador,
    )


def _parse_fecha(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _parse_contador(value: object) -> int | None:
    if value is None or isinstance(value, str) and not value.strip():
        return None
    try:
        return int(float(value))  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None


def _build_device(serie: str, clase: str, items: list[_ParsedRow]) -> DeviceReadings:
    items_sorted = sorted(items, key=lambda i: i.fecha)
    last = items_sorted[-1]
    readings = [CounterReading(i.fecha, i.contador, i.tipo_contador) for i in items_sorted]
    return DeviceReadings(
        serie=serie, clase=clase, articulo=last.articulo, sector=last.sector, readings=readings
    )
